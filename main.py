import os
import psutil
import serial
import shutil
import subprocess
import sys
import time
import warnings
from datetime import datetime, timedelta
from types import SimpleNamespace

import cv2
import pandas as pd
import serial.tools.list_ports as list_ports
from PySide6.QtCore import Qt, QTimer, QTime
from PySide6.QtGui import QTextCursor, QIcon, QImage, QPixmap, QPainter, QPainterPath
from PySide6.QtWidgets import QApplication, QDialog, QMainWindow, QTableWidgetItem, QTableWidget
from gsmmodem.modem import GsmModem
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from sys import *
from ui_continue import Ui_Form
from ui_main import Ui_MainWindow
from ui_notification import Ui_Notification
from ui_settings import Ui_settingswindow


def read_settings():
    settings_file = "Files/settings.txt"
    if not os.path.exists(settings_file):
        print(f"Файл {settings_file} не существует.")
        exit()
    settings = {}
    with open(settings_file, 'r') as file:
        for line_num, line in enumerate(file, start=1):
            line = line.strip()
            if not line or '=' not in line:
                continue
            try:
                name, value = line.split('=', 1)
                settings[name.strip()] = value.strip()
            except ValueError as e:
                print(f"Ошибка в строке {line_num}: '{line}', ошибка: {e}")
                continue
    return settings

def save_settings(settings):
    settings_file = "Files/settings.txt"
    try:
        with open(settings_file, 'w') as file:
            for name, value in settings.items():
                file.write(f"{name}={value}\n")
        print("Настройки успешно сохранены")
    except Exception as e:
        print(f"Ошибка при сохранении настроек: {e}")

class MyDialog(QDialog):
    def __init__(self):
        super(MyDialog, self).__init__()

    def setup(self):
        self.setWindowFlags(Qt.FramelessWindowHint)  # titlebar
        self.setAttribute(Qt.WA_TranslucentBackground)  # пусть фон будет прозрачным чтобы видет видева
        self.setWindowIcon(QIcon("Files/icons/social.ico"))
        self.set_video()

        self.ui.closeButton.clicked.connect(self.close)
        self.ui.minimizeButton.clicked.connect(self.close)


    def set_video(self):
        #настройки
        self.settings = read_settings()
        self.settings['port'] = None  # Настроить
        # инициализация видева

        self.capture = cv2.VideoCapture(f"Files/backgrounds/{self.settings['theme']}.mp4")
        self.snapshot = QPixmap()
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_frame)
        self.start_time = QTime.currentTime()
        # запуск
        self.timer.start(30)

    def update_frame(self):
        if self.capture.isOpened():
            ret, frame = self.capture.read()
            if ret:
                # повтор
                if self.capture.get(cv2.CAP_PROP_POS_FRAMES) == self.capture.get(cv2.CAP_PROP_FRAME_COUNT):
                    self.capture.set(cv2.CAP_PROP_POS_FRAMES, 0)

                frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                h, w, ch = frame.shape
                bytes_per_line = ch * w

                image = QImage(frame.data, w, h, bytes_per_line, QImage.Format_RGB888)
                self.snapshot = QPixmap.fromImage(image)
                self.update()
    def mousePressEvent(self, event):
        try:
            if event.button() == Qt.LeftButton:
                self.dragPos = event.globalPosition().toPoint()
        except:
            pass
    def mouseMoveEvent(self, event):
        try:
            if event.buttons() == Qt.LeftButton:
                self.move(self.pos() + event.globalPosition().toPoint() - self.dragPos)
                self.dragPos = event.globalPosition().toPoint()
                event.accept()
        except:
            pass
    def paintEvent(self, event):
        if not self.snapshot.isNull():
            painter = QPainter(self)
            painter.setRenderHint(QPainter.Antialiasing)

            path = QPainterPath()
            path.addRoundedRect(0, 0, self.width(), self.height(), 10, 10)

            painter.setClipPath(path)

            painter.drawPixmap(0, 0, self.snapshot.scaled(
                self.size(),
                Qt.KeepAspectRatioByExpanding,
                Qt.SmoothTransformation
            ))
        super().paintEvent(event)
    def closeEvent(self, event):
        self.capture.release()
        super().closeEvent(event)
class MyWindow(QMainWindow):
    def __init__(self):
        super(MyWindow, self).__init__()
        self.settings = None
        self.start_time = None

    def setup(self):
        self.center_window()
        self.setWindowIcon(QIcon("Files/icons/social.ico"))

        self.rerun()
        self.incomingCount = 0
        self.settings['port'] = None  # Настроить
        self.canModem = False
        self.setWindowFlags(Qt.FramelessWindowHint)  # titlebar
        self.setAttribute(Qt.WA_TranslucentBackground)  # пусть фон будет прозрачным чтобы видет видева


        self.ui.closeButton.clicked.connect(self.close)
        self.ui.minimizeButton.clicked.connect(self.showMinimized)
    def rerun(self):
        # настройки
        try:
            port = self.settings['port']
            self.settings = read_settings()
            self.settings['port'] = port
        except:
            self.settings = read_settings()


        # инициализация видева
        self.capture = cv2.VideoCapture(f"Files/backgrounds/{self.settings['theme']}.mp4")
        self.snapshot = QPixmap()
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_frame)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.start_time = QTime.currentTime()

        # запуск
        self.timer.start(30)
    def center_window(self):
        screen = self.screen().geometry()
        window = self.geometry()
        self.move(
            screen.center().x() - window.width() // 2,
            screen.center().y() - window.height() // 2
        )
    def update_frame(self):
        if self.capture.isOpened():
            ret, frame = self.capture.read()
            if ret:
                # повтор
                if self.capture.get(cv2.CAP_PROP_POS_FRAMES) == self.capture.get(cv2.CAP_PROP_FRAME_COUNT):
                    self.capture.set(cv2.CAP_PROP_POS_FRAMES, 0)

                frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                h, w, ch = frame.shape
                bytes_per_line = ch * w

                image = QImage(frame.data, w, h, bytes_per_line, QImage.Format_RGB888)
                self.snapshot = QPixmap.fromImage(image)
                self.update()
    def paintEvent(self, event):
        if not self.snapshot.isNull():
            painter = QPainter(self)
            painter.setRenderHint(QPainter.Antialiasing)

            path = QPainterPath()
            path.addRoundedRect(0, 0, self.width(), self.height(), 10, 10)

            painter.setClipPath(path)

            painter.drawPixmap(0, 0, self.snapshot.scaled(
                self.size(),
                Qt.KeepAspectRatioByExpanding,
                Qt.SmoothTransformation
            ))
        super().paintEvent(event)
    def closeEvent(self, event):
        self.capture.release()
        super().closeEvent(event)

class Settings(MyWindow):
    def __init__(self):
        super(Settings, self).__init__()
        self.ui = Ui_settingswindow()
        self.ui.setupUi(self)
        self.setup()
        #настройки
        self.populate_video_combo()

        #значения
        self.ui.modemName.setText(self.settings['model'])
        self.ui.modemName.setCursorPosition(0)
        self.ui.modemName.clearFocus()

        self.ui.modemSpeed.setText(self.settings['speed'])
        self.ui.chooseTheme.setCurrentText(self.settings['theme'])


        # задаём слайдер в настройках
        self.ui.charge_warning.setValue(int(self.settings['charge_warning']))
        self.ui.charge_warning.valueChanged.connect(self.on_charge_warning_change)

        # задаём отображение значения в тексте
        self.ui.charge_warning_display.setText(self.settings['charge_warning'])

        # Привязка кнопок
        self.ui.archivateData.clicked.connect(self.archivate_logs)
        self.ui.cancelSettings.clicked.connect(self.close)
        self.ui.saveSettings.clicked.connect(self.save)

        # Привязка сигнала выбора темы
        self.ui.chooseTheme.currentIndexChanged.connect(self.on_video_changed)
        self.ui.modemSpeed.textChanged.connect(self.updateSpeed)
        self.ui.modemName.textChanged.connect(self.updateModel)

    def updateSpeed(self):
        self.settings['speed'] = self.ui.modemSpeed.text()
    def on_charge_warning_change(self, value):
        self.ui.charge_warning_display.setText(f"{value}%")
        self.settings['charge_warning'] = str(value)
    def updateModel(self):
        self.settings['model'] = self.ui.modemName.text()
    def save(self):
        save_settings(self.settings)

        self.close()
    def mousePressEvent(self, event):
        try:
            if event.button() == Qt.LeftButton and self.childAt(event.pos()) == self.ui.Title:
                self.dragPos = event.globalPosition().toPoint()
        except:
            pass
    def mouseMoveEvent(self, event):
        try:
            if event.buttons() == Qt.LeftButton and self.childAt(event.pos()) == self.ui.Title:
                self.move(self.pos() + event.globalPosition().toPoint() - self.dragPos)
                self.dragPos = event.globalPosition().toPoint()
                event.accept()
        except:
            pass

    def archivate_logs(self):
        log_file = "Files/sms_log.xlsx"
        analysis_file = "Files/Analysis.txt"
        archive_dir = "Files/Archive"
        if not os.path.exists(archive_dir):
            os.makedirs(archive_dir)

        current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        archived_log_file = f"{archive_dir}/sms_log_{current_datetime}.xlsx"
        archived_analysis_file = f"{archive_dir}/Analysis_{current_datetime}.txt"

        try:
            shutil.copy2(log_file, archived_log_file)
            print(f"Лог успешно архивирован в {archived_log_file}")

            wb = load_workbook("Files/sms_log.xlsx", data_only=True)
            ws = wb.active
            # Удаляем все строки, кроме первой
            ws.delete_rows(2, ws.max_row)

            # Устанавливаем высоту первой строки по умолчанию
            ws.row_dimensions[1].height = 13.7

            # Удаляем все настройки высоты строк, чтобы они стали по умолчанию
            for row_dim in ws.row_dimensions:
                if row_dim != 1:
                    ws.row_dimensions[row_dim].height = None

            wb.save(log_file)

            # Архивирование файла analysis.txt
            shutil.copy2(analysis_file, archived_analysis_file)
            print(f"Файл analysis.txt успешно архивирован в {archived_analysis_file}")

            # Очистка файла analysis.txt
            with open(analysis_file, 'w') as f:
                f.write('')

            print("Файл analysis.txt успешно очищен")

        except Exception as e:
            print(f"Ошибка при очистке логов: {e}")
    def populate_video_combo(self):
        # Путь к папке с фонами
        backgrounds_path = "Files/backgrounds"

        # Создаем папку если её нет
        if not os.path.exists(backgrounds_path):
            os.makedirs(backgrounds_path)

        # Получаем список всех .mp4 файлов
        video_files = [f for f in os.listdir(backgrounds_path) if f.endswith('.mp4')]

        # Добавляем имена файлов без расширения в комбобокс
        for video in video_files:
            name_without_ext = os.path.splitext(video)[0]
            # Сохраняем полный путь в userData для последующего использования
            self.ui.chooseTheme.addItem(name_without_ext, userData=os.path.join(backgrounds_path, video))
    def on_video_changed(self, index):
        self.settings['theme'] = self.ui.chooseTheme.currentText()
        video_path = self.ui.chooseTheme.itemData(index)
        self.capture = cv2.VideoCapture(video_path)


class Continue(MyDialog):
    def __init__(self, text, but1text, but2text):
        super(Continue, self).__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        self.setup()

        self.choice = None  #выбор пользователя

        self.ui.but1.setText(but1text)
        self.ui.but2.setText(but2text)
        self.ui.text.setText(text)

        self.ui.but1.clicked.connect(self.but1)
        self.ui.but2.clicked.connect(self.but2)

    def but1(self):
        self.choice = False  # Запоминаем выбор
        self.accept()#0
    def but2(self):
        self.choice = True  # Запоминаем выбор
        self.accept()#1
    def get_choice(self):
        if self.exec() == QDialog.Accepted:
            return self.choice
        return None


class Notification(MyDialog):
    def __init__(self, text):
        super(Notification, self).__init__()
        self.ui = Ui_Notification()
        self.ui.setupUi(self)

        self.setup()

        self.ui.but2.setText("OK")
        self.ui.text.setText(text)

        self.ui.but2.clicked.connect(self.but2)

    def but2(self):
        self.close()
    def run(self):
        if self.exec() == QDialog.Accepted:
            return self.choice
        return None

class SmsTools(MyWindow):
    def __init__(self):
        super(SmsTools, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.setup()
        self.ui.centralwidget.setStyleSheet("""
            background: transparent;
            border-radius: 20px;
        """)
        self.files = SimpleNamespace()
        self.filePaths = SimpleNamespace()
        self.contacts = None
        self.modem = None

        # биндинг кнопок
        self.ui.rerunButton.clicked.connect(self.rerun)

        self.ui.sendButton.clicked.connect(self.send)
        self.ui.restart.clicked.connect(self.restart_modem)
        self.ui.get_messages.clicked.connect(self.read_sms_and_save)
        self.ui.add_contact.clicked.connect(self.add_contacts)
        self.ui.analyze.clicked.connect(self.menu_analysis)
        self.ui.open_folder.clicked.connect(self.open_files_folder)
        self.ui.save.clicked.connect(self.edit_contacts)
        self.ui.search.clicked.connect(self.load_contacts)
        self.ui.settings.clicked.connect(self.openSettings)

        self.load_contacts(search_terms="0 1 2 3 4 5 6 7 8 9")
        self.kill_connect_manager()
        self.setup_modem()
    def mousePressEvent(self, event):
        try:
            if event.button() == Qt.LeftButton:
                self.dragPos = event.globalPosition().toPoint()
        except:
            pass
    def mouseMoveEvent(self, event):
        try:
            if event.buttons() == Qt.LeftButton:
                self.move(self.pos() + event.globalPosition().toPoint() - self.dragPos)
                self.dragPos = event.globalPosition().toPoint()
                event.accept()
        except:
            pass
    def add_contacts(self):
        number = self.ui.number.text()
        name = self.ui.name.text()

        contact = [number, name]
        file_path = self.filePaths.contacts
        directory = os.path.dirname(file_path)
        if not os.path.exists(directory):
            os.makedirs(directory)

        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Contacts"
            # заголовки
            ws.append(["Номер телефона", "Имя маячка"])

        ws.append(contact)

        wb.save(file_path)

        self.load_contacts()
        print(f"Контакты успешно добавлены в {file_path}")
        return None
    def setup_modem(self):
        #настройка файлов

        self.filePaths.smsLog = "Files/sms_log.xlsx"
        self.filePaths.contacts = "Files/contacts.xlsx"
        self.files.smsLog = load_workbook(self.filePaths.smsLog, data_only=True)  # Загружаем
        self.files.contacts = load_workbook(self.filePaths.contacts, data_only=True)  # Загружаем
        self.contacts = self.get_all_contacts()  # Загружаем

        #Инициализация модема
        self.kill_connect_manager()  # Отключение коннект-менеджера
        self.find_modem()  # Попытка найти модем
        if self.canModem:  # Если модем найден, подключение. ОБязательные вещи.
            print(self.send_at_command('AT+CMGF=1'))  # Задание текстового формата
            self.send_at_command('AT+CPMS="ME","ME","ME"')  # переключение на внутреннюю память
            # self.conprint(self.modem.read())
    # функция для объединения сообщений
    def combine_long_messages(self, messages):
        combined_messages = []
        for message in messages:
            combined_messages.append(message)
        return combined_messages
    def send_at_command(self, command, timeout=0.1):
        if not self.canModem:
            Notification("Модем не подключен.").run()
            return
        with serial.Serial(self.settings['port'], self.settings['speed'], timeout=timeout) as ser:
            ser.write((command + '\r\n').encode())
            time.sleep(timeout)
            response = ser.read_all().decode()
            return response
    def find_modem(self):
        # Находим все доступные порты
        available_ports = list_ports.comports()

        if not available_ports:
            print("Не удалось найти модем.")
            self.conprint("Не удалось найти модем.")
            print('Функции отправки и принятия СМС не будут работать.')
            self.conprint('Функции отправки и принятия СМС не будут работать.')
        else:
            # проходим по каждому доступному порту
            for port_info in available_ports:
                port = port_info.device
                device_name = port_info.description  #получаем имя устройства
                if self.settings['model'] in device_name:
                    try:
                        ser = serial.Serial(port, timeout=2)
                        ser.write(b'AT\r\n')
                        response = ser.read(100).decode('utf-8').strip()
                        ser.close()

                        if response:
                            # сохраняем первый найденный порт и завершаем
                            self.settings['port'] = port
                            self.conprint("Модем подключен.")
                            self.canModem = True
                            break
                    except serial.SerialException as e:
                        print(self.conprint(f'ОШИБКА: {e}'))
    def num_to_name(self, num):
        # преобразует номер в имя
        wb = self.files.contacts
        ws = wb.active
        # print(f"Искомый номер: {num}")
        for row in ws.iter_rows(min_row=2, values_only=True):
            phone_number, contact_name = row
            if phone_number:
                phone_number = f"+7{phone_number}".replace(" ", "")
                # print(f"Номер контакта: {phone_number}")

                if phone_number == num:
                    return contact_name
        return num
    def get_current_datetime(self):
        now = datetime.now()
        return now.strftime('%d/%m/%Y'), now.strftime('%H:%M:%S')
    # Функция для парсинга сообщений
    def parse_sms_response(self, response):
        messages = []
        lines = response.splitlines()
        i = 0
        '''
        пример сообщения: 
        +CMGL: 10,"REC READ","+79875324724",,"24/11/15,14:35:51+12"
        Hello!
        Или:
        +CMGL: 0,"REC UNREAD","+79302590039",,"25/02/25,16:41:11+12"
        '''
        while i < len(lines):
            if "+CMGL: " in lines[i]:
                parts = lines[i].split(",")
                index = parts[0].split(": ")[1].strip()
                sender_number = parts[2].strip('"')
                date_and_time = lines[i].split(",,")[1].replace('"', '').split(',')
                # print(f"{date_and_time=}")
                date_dates = date_and_time[0].split("/")
                date_date = f"{date_dates[2]}.{date_dates[1]}.{date_dates[0]}"
                # print(f"ДАТА = {date_date}")

                date_time = date_and_time[1].split("+")[0].split("-")[0]
                # print(f"ВРЕМЯ = {date_time}")

                message_lines = []
                if i + 1 < len(lines):
                    j = i + 1
                    while j < len(lines) and "+CMGL: " not in lines[j]:
                        if "OK" in lines[j]:
                            break
                        message_lines.append(lines[j].strip())
                        j += 1
                    i = j - 1

                decoded_lines = []
                for line in message_lines:
                    try:
                        decoded_line = bytes.fromhex(line).decode('utf-16be')
                        decoded_lines.append(decoded_line)
                    except (ValueError, UnicodeDecodeError):
                        decoded_lines = message_lines
                        break

                message = '\n'.join(decoded_lines)

                messages.append({
                    "index": index,
                    "sender_number": sender_number,
                    "date": date_date,
                    "time": date_time,
                    "message": message.strip()
                })
            i += 1
        return messages
    def get_all_contacts(self):
        try:
            df = pd.read_excel(self.filePaths.contacts)
        except Exception as e:
            print(f"Ошибка при чтении файла контактов: {e}")
            return {}

        required_columns = ['Номер телефона', 'Имя маячка']
        for column in required_columns:
            if column not in df.columns:
                print(f"Отсутствует ожидаемый столбец: '{column}'")
                return {}

        contacts = {}
        for index, row in df.iterrows():
            # print(row['Номер телефона'])
            # приведение номеров телефонов к строковому формату без лишних символов
            phone_number = str(row['Номер телефона']).replace(' ', '').replace('-', '')
            contacts[phone_number] = row['Имя маячка']
        return contacts
    def read_sms_and_save(self):
        if not self.canModem:
            Notification("Модем не подключен.").run()
            return
        # print("Проверяем...")
        response = self.send_at_command('AT+CMGL="ALL"')

        #обработка ответа и запись в таблицу
        sms_messages = self.parse_sms_response(response)
        combined_messages = self.combine_long_messages(sms_messages)

        if not os.path.exists(str(self.filePaths.contacts)):
            self.conprint(f"Файл контактов не найден.")
            return
        # self.conprint("Файл контактов найден")
        if combined_messages:
            latest_name = "test4d!"
            log = ""
            for sms in combined_messages:
                # print('')
                name = self.num_to_name(sms['sender_number'])
                if name != latest_name:
                    print(f">> {self.num_to_name(sms['sender_number'])}: \n{sms['message']}")
                    self.conprint(f">> {self.num_to_name(sms['sender_number'])}: \n{sms['message']}")
                else:
                    print(f"{sms['message']}")
                    self.conprint(f"{sms['message']}")
                latest_name = name
            self.append_to_excel(combined_messages)
            # print("Добавлено, удаляем")
            # Удаление
            # по индексу
            for sms in combined_messages:
                # print(f"удаляем {sms}")
                self.send_at_command(f"AT+CMGD={sms['index']}")
                self.incomingCount = 0
            return log
        else:
            cy = 1
            if cy == 15:
                cy = 1
            self.incomingCount = 0
            self.ui.get_messages.setText("Получить сообщения")
            return ""
    def append_to_excel(self, sms_messages):
        if not sms_messages:
            return
        wb = self.files.smsLog
        ws = wb.active

        settings = self.settings
        sleep_time = int(settings.get('sleep_time', 0))

        for sms in sms_messages:
            sender_number = sms["sender_number"].replace(' ', '').replace('-',
                                                                          '')
            contact_name = self.num_to_name(sms['sender_number'])
            message = sms["message"] if sms["message"] else "Без текста"
            date_received = sms["date"]
            current_date = datetime.now().strftime('%d/%m/%Y')
            current_time = datetime.now().strftime('%H:%M:%S')

            existing_row = None
            for row in ws.iter_rows(min_row=2, values_only=False):
                if (row[0].value == sender_number and
                        row[3].value == date_received and
                        abs((datetime.strptime(current_time, '%H:%M:%S') - datetime.strptime(row[4].value,
                                                                                             '%H:%M:%S')).total_seconds()) <= sleep_time + 30):
                    existing_row = row
                    break

            if existing_row:
                existing_row[2].value += "\n" + message
                lines = existing_row[2].value.count('\n') + 1
                ws.row_dimensions[existing_row[0].row].height = 13.7
            else:
                ws.append([sender_number, contact_name, message, date_received, current_time])
                lines = message.count('\n') + 1
                ws.row_dimensions[ws.max_row].height = 13.7

        wb.save(self.filePaths.smsLog)
        self.files.smsLog = wb
    def restart_modem(self, text="Перезагрузить модем?"):
        if Continue(f"{text}\nЗаймёт 50 секунд.", "Нет", "Да").get_choice():
            res = self.send_at_command('AT+CFUN=1,1')

            time.sleep(50)
            self.setup_modem()
            Notification("Модем перезагружен").run()
            return True if "OK" in res else False
    def kill_connect_manager(self):
        try:
            for proc in psutil.process_iter(['name']):
                if proc.info['name'] and 'Connect Manager.exe' in proc.info['name']:
                    proc.kill()
                    # print("Процесс успешно завершен")
                    return True

            # print("Процесс Connect Manager.exe не найден")
            return False

        except Exception as e:
            print(f"Произошла ошибка: {e}")
            return False
    def menu_analysis(self):
        if Continue("Анализировать данные?", "Нет", "Да").get_choice():
            print("Анализирую...")
            self.analyze_smsLog()
            Notification("Анализ завершён.").run()
            print("Конец анализа.")
    def openSettings(self):  # Открывает настройки
        self.settingsWindow = Settings()
        self.settingsWindow.show()
    def delete_contact(self, nums):
        ii = 0
        try:
            # Загружаем файл
            wb = load_workbook("Files/contacts.xlsx")
            ws = wb.active
            # находим все для удаления в обратном порядке
            rows_to_delete = []
            for row in range(ws.max_row, 1, -1):  # начинаем с конца, пропускаем заголовок
                if f"+7{ws.cell(row=row, column=1).value}" in nums:
                    ii += 1
                    rows_to_delete.append(row)

            # удаляем
            for row_idx in rows_to_delete:
                ws.delete_rows(row_idx, 1)

            # Сохраняем
            wb.save("Files/contacts.xlsx")
            return True, f"{ii} Контактов успешно удалено"

        except Exception as e:
            return False, f"Ошибка при удалении контакта: {str(e)}"
    def check_sms_symbols(self, message):
        # Проверяет текст на наличие недопустимых символов для текстового режима
        allowed_chars = set(
            'abcdefghijklmnopqrstuvwxyz'
            'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            '0123456789'
            ' .,!?()-+=:;@')

        for char in message:
            if char not in allowed_chars:
                return False
        return True
    def send(self):
        if not self.canModem:
            Notification("Модем не подключен.").run()
            return
        # try:
        if True:
            modem = GsmModem(self.settings['port'], self.settings['speed'])
            modem.connect("")
            message = self.ui.message.toPlainText()

            if not message:
                Notification("Введите сообщение!").run()
                return

            recipient_numbers = self.get_selected_numbers()
            if not recipient_numbers:
                Notification("Выберите контакты!").run()
                return

            use_text_mode = self.check_sms_symbols(message)  # use PDU mode
            self.conprint(f"\nОтправка сообщений ({len(recipient_numbers)})...")
            print("\nОтправка сообщений\n[", end="")

            modem.smsTextMode = use_text_mode

            # self.modem.connect("")
            for recipient_number in recipient_numbers:
                modem.sendSms(recipient_number, message)
                print("#", end="")
            print("]")
            self.conprint("Сообщения отправлены.")
            modem.close()
            modem.connect("")
            modem.smsTextMode = True
            modem.close()
        # except:
        #     self.restart_modem(text="Не удалось отправить сообщения, перезагрузить модем?")

    def get_selected_numbers(self):
        selected_rows = set()
        for index in self.ui.contacts.selectedIndexes():
            selected_rows.add(index.row())

        first_column_data = []
        for row in selected_rows:
            item = self.ui.contacts.item(row, 0)
            if item is not None:
                first_column_data.append(item.text())

        # self.conprint(f"Выделено {(first_column_data)}")  # Выводим в консоль (можно использовать иначе)
        return first_column_data
    def open_files_folder(self):
        # Открывает папку 'Files' в текущем каталоге в проводнике.

        try:
            # Определяем путь к папке 'Files'
            current_directory = os.getcwd()
            folder_path = os.path.join(current_directory, 'Files')

            # Проверяем, существует ли папка
            if not os.path.isdir(folder_path):
                print(f"Папка не существует: {folder_path}")
                return

            # Открываем папку в проводнике
            if os.name == 'nt':  # Windows
                subprocess.run(['explorer', folder_path], check=True)
            elif os.name == 'posix':  # macOS or Linux
                if sys.platform == 'darwin':  # macOS
                    subprocess.run(['open', folder_path], check=True)
                else:  # Linux
                    subprocess.run(['xdg-open', folder_path], check=True)
            else:
                print(f"Операционная система не поддерживается: {os.name}")

        except subprocess.CalledProcessError as e:
            print('', end='')
        except Exception as e:
            print('', end='')
    def load_smsLog(self, filename):
        try:
            df = pd.read_excel(filename)
            df = df.drop_duplicates()
            df['Сообщение'] = df['Сообщение'].astype(str)
            # print("Загруженные столбцы:", df.columns)
        except Exception as e:
            print(f"Ошибка при чтении файла SMS логов: {e}")
            return pd.DataFrame()
        return df
    def analyze_smsLog(self):
        smsLog = self.load_smsLog(self.filePaths.smsLog)

        if smsLog.empty:
            print(f"Список сообщений пуст.")
            return
        # Получение даты и времени
        today_date, current_time = self.get_current_datetime()
        yesterday_date = (datetime.now() - timedelta(days=1)).strftime('%d/%m/%Y')

        # Используем индексаци
        phone_column_index = 0  # Индекс столбца с номерами телефонов
        date_column_index = 3  # Индекс столбца с датой получения SMS

        # Проверка типов
        smsLog.iloc[:, phone_column_index] = smsLog.iloc[:, phone_column_index].astype(str)
        smsLog.iloc[:, date_column_index] = pd.to_datetime(smsLog.iloc[:, date_column_index], format='%d/%m/%Y',
                                                           errors='coerce').dt.strftime('%d/%m/%Y')

        # Анализ сообщений
        settings = self.settings
        charge_warning = int(settings.get('charge_warning', 0))
        wb = self.files.smsLog
        ws = wb.active

        # Очистка 7-го столбца
        for row in ws.iter_rows(min_row=2, values_only=False):
            row[5].value = ""

        # Добавляем столбец "Отклонения", если его нет
        if ws.max_column < 6:
            ws['G1'] = 'Отклонения'

        for row in ws.iter_rows(min_row=2, values_only=False):
            message = row[2].value
            if isinstance(message, str):  # Проверяем, является ли сообщение строкой
                deviations = row[5].value if row[5].value else ""
                battery_warning = False
                gps_warning = False
                for line in message.splitlines():
                    if "Спутн: 0" in line:
                        gps_warning = True
                    if "Бат:" in line:
                        battery_level = int(line.split("(")[1].split("%")[0])
                        if battery_level < charge_warning:
                            battery_warning = True
                if battery_warning:
                    deviations += "Бат! "
                if gps_warning:
                    deviations += "GPS! "
                row[5].value = deviations

                # Устанавливаем цвет фона для 7-го столбца
                if "Бат! GPS! " in deviations:
                    row[5].fill = PatternFill(start_color='FFFF950E', end_color='FFFF950E', fill_type='solid')
                elif "GPS! " in deviations:
                    row[5].fill = PatternFill(start_color='FFF0F076', end_color='FFF0F076', fill_type='solid')
                elif "Бат! " in deviations:
                    row[5].fill = PatternFill(start_color='FFAFEEEE', end_color='FFAFEEEE', fill_type='solid')

        # Установка белой заливки для пустых ячеек в 7-м столбце
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[5].value is None or row[5].value == "":
                row[5].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        wb.save(self.filePaths.smsLog)
    def keyPressEvent(self, event):
        print(event)
        # чтобы нажать esq и например отменить выделение или нажать enter и отправить сообщение (потом)
        if self.ui.contacts.hasFocus():
            if event.key() == Qt.Key_Escape:
                print("Очистка.")
                self.ui.contacts.clearSelection()
            elif event.key() == Qt.Key_Backspace:
                numbers = self.get_selected_numbers()
                if Continue(f"Удалить выделенные контакты ({len(numbers)})?", "Нет", "Да").get_choice():
                    self.delete_contact(numbers)
                    self.ui.contacts.clearSelection()
                    self.load_contacts()
        elif self.ui.name.hasFocus() or self.ui.number.hasFocus():
            if event.key() == Qt.Key_Return:
                self.add_contacts()
        elif self.ui.arguments.hasFocus():
            if event.key() == Qt.Key_Return:
                self.load_contacts()
    def search_contacts(self, file_path, search_terms):
        wb = load_workbook(file_path)
        ws = wb.active

        if search_terms == "":
            search_terms = ["9", "8", "7", "6", "5", "4", "3", "2", "1", "0"]
        else:
            search_terms = search_terms.split()

        final_strings = []

        for search_term in search_terms:
            if not search_term.startswith("-"):
                for row in ws.iter_rows(min_row=2, values_only=True):
                    phone_number, contact_name = row
                    if phone_number:
                        string = f"{phone_number}::{contact_name}".replace(" ", "")
                        if search_term in string and string not in final_strings:
                            final_strings.append(string)
            else:
                argument = search_term.replace("-", "")
                for final_string in final_strings[:]:  # Используем копию списка
                    if argument in final_string:
                        final_strings.remove(final_string)

        contacts_found = [{"num": f"+7{num}", "name": name} for num, name in
                          (s.split("::") for s in final_strings)]

        if not contacts_found:
            return []

        return contacts_found
    def load_contacts(self, search_terms=None):
        file_path = "Files/contacts.xlsx"
        if not search_terms:
            search_terms = self.ui.arguments.text()
        contacts = self.search_contacts(file_path, search_terms)  # Получаем список контактов

        self.populate_table(contacts)
        return contacts
    def edit_contacts(self):
        if Continue(f"Сохранить контакты?", "Нет", "Да").get_choice():
            edited_contacts = self.get_displayed_contacts()
            file_path = "Files/contacts.xlsx"
            all_contacts = self.search_contacts(file_path, "0 1 2 3 4 5 6 7 8 9")  # Получаем список контактов

            # Создаем словарь, где ключ — это номер телефона, а значение — новое имя
            edited_dict = {contact['num']: contact['name'] for contact in edited_contacts}

            # Обновляем all_contacts, если номер найден в edited_dict
            for contact in all_contacts:
                if contact['num'] in edited_dict:
                    contact['name'] = edited_dict[contact['num']]
            # print(edited_contacts)

            self.rewrite_contacts(all_contacts)
            # self.load_contacts()
    def rewrite_contacts(self, contacts):
        file_path = "Files/contacts.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Контакты"
        ws.append(["Номер телефона", "Имя маячка"])  # Заголовки

        for contact in contacts:
            ws.append([contact["num"].replace("+7", ""), contact["name"]])

        wb.save(file_path)
    def get_displayed_contacts(self):
        contacts = []
        for row in range(self.ui.contacts.rowCount()):
            num_item = self.ui.contacts.item(row, 0)  # Получаем номер
            name_item = self.ui.contacts.item(row, 1)  # Получаем имя

            if num_item and name_item:  # Проверяем, что ячейки не пустые
                contacts.append({"num": num_item.text(), "name": name_item.text()})

        return contacts
    def populate_table(self, contacts):
        self.ui.contacts.setRowCount(len(contacts))
        self.ui.contacts.setColumnCount(2)
        self.ui.contacts.setHorizontalHeaderLabels(["Номер", "Имя"])

        self.ui.contacts.setColumnWidth(0, 100)  # Номер телефона
        if len(contacts) <= 7:
            self.ui.contacts.setColumnWidth(1, 89)  # больше
        else:
            self.ui.contacts.setColumnWidth(1, 78)  # меньше
        self.ui.contacts.setShowGrid(False)


        self.ui.contacts.verticalHeader().setDefaultSectionSize(30)  # Высота строк

        self.ui.contacts.setSelectionBehavior(QTableWidget.SelectRows)

        self.ui.contacts.horizontalHeader().hide()
        self.ui.contacts.verticalHeader().hide()

        for row, contact in enumerate(contacts):
            # заполнение
            phone_item = QTableWidgetItem(contact["num"])
            phone_item.setFlags(phone_item.flags() & ~Qt.ItemIsEditable)  # Убираем флаг редактирования
            self.ui.contacts.setItem(row, 0, phone_item)

            name_item = QTableWidgetItem(contact["name"])
            self.ui.contacts.setItem(row, 1, name_item)
    def conprint(self, text):
        self.ui.console.setText(f"{self.ui.console.toPlainText()}\n{text}")
        self.ui.console.moveCursor(QTextCursor.MoveOperation.End)
        self.ui.console.ensureCursorVisible()  # прокрутка


if __name__ == "__main__":
    # print(edit_contacts())
    warnings.simplefilter(action='ignore', category=FutureWarning)
    app = QApplication(sys.argv)

    window = SmsTools()
    window.show()


    sys.exit(app.exec())
