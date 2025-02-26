import sys, os, subprocess, shutil
from datetime import datetime
from PySide6.QtCore import Qt, QTimer, QTime
from PySide6.QtWidgets import QApplication, QMainWindow, QTableWidgetItem, QTableWidget, QTextBrowser, QTextEdit, \
    QPushButton, QVBoxLayout, QWidget, QComboBox
from Scripts.bottle import delete
from openpyxl import load_workbook, Workbook
from ui_settings import Ui_settingswindow
import cv2
from PySide6.QtGui import QImage, QPixmap, QPainter, QPainterPath


def read_settings():
    settings_file = "Files/settings.txt"
    if not os.path.exists(settings_file):
        print(f"Файл {settings_file} не существует.")
        exit()
    settings = {}
    with open(settings_file, 'r') as file:
        for line_num, line in enumerate(file, start=1):
            line = line.strip()
            if not line or '=' not in line:
                continue
            try:
                name, value = line.split('=', 1)
                settings[name.strip()] = value.strip()
            except ValueError as e:
                print(f"Ошибка в строке {line_num}: '{line}', ошибка: {e}")
                continue
    return settings

def save_settings(settings):
    settings_file = "Files/settings.txt"
    try:
        with open(settings_file, 'w') as file:
            for name, value in settings.items():
                file.write(f"{name}={value}\n")
        print("Настройки успешно сохранены")
    except Exception as e:
        print(f"Ошибка при сохранении настроек: {e}")


class Settings(QMainWindow):
    def __init__(self):
        super(Settings, self).__init__()
        self.ui = Ui_settingswindow()
        self.ui.setupUi(self)

        #настройки
        self.settings = read_settings()
        self.populate_video_combo()

        #значения
        self.ui.modemName.setText(self.settings['model'])
        self.ui.modemName.setCursorPosition(0)
        self.ui.modemName.clearFocus()
        self.ui.modemName.clearFocus()

        self.ui.modemSpeed.setText(self.settings['speed'])
        self.ui.Animated.setChecked(bool(self.settings['animated']))
        self.ui.chooseTheme.setCurrentText(self.settings['theme'])

        # Инициализация видео
        self.capture = cv2.VideoCapture(f'{self.settings["theme"]}.mp4')
        self.snapshot = QPixmap()
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_frame)
        self.start_time = QTime.currentTime()

        # Запускаем воспроизведение
        self.timer.start(30)  # 30ms = ~33fps

        # Делаем фон прозрачным
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setWindowFlags(Qt.FramelessWindowHint)  # titlebar

        # Привязка кнопок
        self.ui.archivateData.clicked.connect(self.archivate_logs)
        self.ui.cancelSettings.clicked.connect(self.cancelSettings)
        self.ui.closeButton.clicked.connect(self.close)
        self.ui.minimizeButton.clicked.connect(self.showMinimized)

        # Привязка сигнала выбора темы
        self.ui.chooseTheme.currentIndexChanged.connect(self.on_video_changed)


    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.dragPos = event.globalPosition().toPoint()

    def update_frame(self):
        if self.capture.isOpened():
            ret, frame = self.capture.read()
            if ret:
                # Если достигнут конец видео, начинаем сначала
                if self.capture.get(cv2.CAP_PROP_POS_FRAMES) == self.capture.get(cv2.CAP_PROP_FRAME_COUNT):
                    self.capture.set(cv2.CAP_PROP_POS_FRAMES, 0)

                # Конвертируем BGR в RGB
                frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                h, w, ch = frame.shape
                bytes_per_line = ch * w

                # Создаем QImage из кадра
                image = QImage(frame.data, w, h, bytes_per_line, QImage.Format_RGB888)
                self.snapshot = QPixmap.fromImage(image)
                self.update()  # Вызываем перерисовку

    def paintEvent(self, event):
        if not self.snapshot.isNull():
            painter = QPainter(self)
            painter.setRenderHint(QPainter.Antialiasing)

            path = QPainterPath()
            path.addRoundedRect(0, 0, self.width(), self.height(), 10, 10)

            painter.setClipPath(path)

            painter.drawPixmap(0, 0, self.snapshot.scaled(
                self.size(),
                Qt.KeepAspectRatioByExpanding,
                Qt.SmoothTransformation
            ))
        super().paintEvent(event)

    def closeEvent(self, event):
        self.capture.release()
        super().closeEvent(event)

    def keyPressEvent(self, event):
        pass

    def archivate_logs(self):
        log_file = "Files/sms_log.xlsx"
        analysis_file = "Files/Analysis.txt"
        archive_dir = "Files/Archive"
        if not os.path.exists(archive_dir):
            os.makedirs(archive_dir)

        current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        archived_log_file = f"{archive_dir}/sms_log_{current_datetime}.xlsx"
        archived_analysis_file = f"{archive_dir}/Analysis_{current_datetime}.txt"

        try:
            shutil.copy2(log_file, archived_log_file)
            print(f"Лог успешно архивирован в {archived_log_file}")

            wb = load_workbook("Files/sms_log.xlsx", data_only=True)
            ws = wb.active
            # Удаляем все строки, кроме первой
            ws.delete_rows(2, ws.max_row)

            # Устанавливаем высоту первой строки по умолчанию
            ws.row_dimensions[1].height = 13.7

            # Удаляем все настройки высоты строк, чтобы они стали по умолчанию
            for row_dim in ws.row_dimensions:
                if row_dim != 1:
                    ws.row_dimensions[row_dim].height = None

            wb.save(log_file)
            print("Лог успешно очищен, кроме строки заголовков")

            # Архивирование файла analysis.txt
            shutil.copy2(analysis_file, archived_analysis_file)
            print(f"Файл analysis.txt успешно архивирован в {archived_analysis_file}")

            # Очистка файла analysis.txt
            with open(analysis_file, 'w') as f:
                f.write('')

            print("Файл analysis.txt успешно очищен")

        except Exception as e:
            print(f"Ошибка при очистке логов: {e}")

    def cancelSettings(self):
        self.close()

    def populate_video_combo(self):
        # Путь к папке с фонами
        backgrounds_path = "Files/backgrounds"

        # Создаем папку если её нет
        if not os.path.exists(backgrounds_path):
            os.makedirs(backgrounds_path)
        print("получаем")
        # Получаем список всех .mp4 файлов
        video_files = [f for f in os.listdir(backgrounds_path) if f.endswith('.mp4')]

        # Добавляем имена файлов без расширения в комбобокс
        for video in video_files:
            name_without_ext = os.path.splitext(video)[0]
            # Сохраняем полный путь в userData для последующего использования
            self.ui.chooseTheme.addItem(name_without_ext, userData=os.path.join(backgrounds_path, video))

    def on_video_changed(self, index):
        # Получаем полный путь к файлу из userData
        video_path = self.ui.chooseTheme.itemData(index)
        # Здесь можно обновить видео
        self.capture = cv2.VideoCapture(video_path)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = Settings()
    window.show()
    sys.exit(app.exec())